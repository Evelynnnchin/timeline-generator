import re
from io import BytesIO
from typing import Dict, List, Tuple

import openpyxl
import pandas as pd
import plotly.express as px
import streamlit as st


# =========================
# PAGE SETUP
# =========================
st.set_page_config(page_title="T&C Resource Timeline", layout="wide")
st.title("📊 T&C Resource Gantt / Overload Checker")
st.caption(
    "Upload the resource planning workbook. The app reads only the `T&C Activities` sheet, "
    "then flags any role/month where the total manpower loading exceeds 1."
)


# =========================
# CONFIG
# =========================
SHEET_NAME = "T&C Activities"
MONTH_MAP: Dict[str, int] = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}

# These rows are category/summary rows in the T&C Activities sheet, not individual people.
SUMMARY_ROLES = {
    "TOTAL",
    "ATC/ATS",
    "ATS",
    "SIG",
    "ATC",
    "COMMS",
    "SUBCON",
    "TRAIN",
    "CSF",
}


# =========================
# HELPERS
# =========================
def _is_number_like(value) -> bool:
    """Return True if value looks like an activity serial number, e.g. 1, 2, 10.1."""
    if value is None:
        return False
    return bool(re.fullmatch(r"\d+(\.\d+)?", str(value).strip()))


def _clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


@st.cache_data(show_spinner=False)
def parse_tc_activities(file_bytes: bytes) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Parse only the T&C Activities sheet.

    Expected workbook structure:
    - Row 1: years across the timeline, usually merged visually in Excel
    - Row 2: month names
    - Column A: project section or activity number
    - Column B: activity name or role name
    - Column D: requirement text
    - Month columns: manpower loading values
    """
    wb = openpyxl.load_workbook(
        filename=BytesIO(file_bytes),
        read_only=True,
        data_only=True,
        keep_links=False,
    )

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Cannot find sheet named `{SHEET_NAME}`. Available sheets: {', '.join(wb.sheetnames)}"
        )

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 3:
        raise ValueError("The T&C Activities sheet does not have enough rows to parse.")

    max_col = max(len(row) for row in rows)
    row_1 = rows[0]
    row_2 = rows[1]

    # Build monthly timeline columns from row 1 / row 2.
    # Row 1 only has the year value at the first month of each year, so we forward-fill it.
    month_cols: List[Tuple[int, pd.Timestamp]] = []
    current_year = None

    for col_idx in range(max_col):
        year_value = row_1[col_idx] if col_idx < len(row_1) else None
        month_value = row_2[col_idx] if col_idx < len(row_2) else None

        if isinstance(year_value, (int, float)) and int(year_value) >= 2000:
            current_year = int(year_value)

        if isinstance(month_value, str):
            month_key = month_value.strip().upper()[:3]
            if month_key in MONTH_MAP and current_year is not None:
                month_cols.append(
                    (col_idx, pd.Timestamp(current_year, MONTH_MAP[month_key], 1))
                )

    if not month_cols:
        raise ValueError("No timeline month columns were found in row 1 / row 2.")

    records = []
    current_project = ""
    current_activity_no = ""
    current_activity = ""
    current_requirement = ""

    for excel_row_no, row in enumerate(rows[2:], start=3):
        def get_cell(idx: int):
            return row[idx] if idx < len(row) else None

        col_a = get_cell(0)
        col_b = get_cell(1)
        col_d = get_cell(3)

        a_text = _clean_text(col_a)
        b_text = _clean_text(col_b)
        d_text = _clean_text(col_d)

        if not a_text and not b_text:
            continue

        # Project section rows, e.g. DTL / JRL / CRL / RTS.
        if a_text and not b_text and not _is_number_like(a_text):
            current_project = a_text
            continue

        # Activity rows: serial number in Column A + activity name in Column B.
        if a_text and b_text and _is_number_like(a_text):
            current_activity_no = a_text
            current_activity = b_text
            current_requirement = d_text
            continue

        # Role rows: role name in Column B + manpower values in monthly columns.
        if not a_text and b_text:
            role = b_text
            for col_idx, month_start in month_cols:
                load_value = get_cell(col_idx)

                if isinstance(load_value, (int, float)) and float(load_value) != 0:
                    records.append(
                        {
                            "Project": current_project,
                            "Activity No": current_activity_no,
                            "Activity": current_activity,
                            "Requirement": current_requirement,
                            "Role": role,
                            "Month": month_start,
                            "Load": float(load_value),
                            "Source Row": excel_row_no,
                        }
                    )

    if not records:
        raise ValueError("No role/month manpower values were found.")

    detail_df = pd.DataFrame(records)
    detail_df["Month Label"] = detail_df["Month"].dt.strftime("%b-%Y")

    # Total monthly loading by role. This is the main overload check.
    role_month_df = (
        detail_df.groupby(["Role", "Month"], as_index=False)["Load"]
        .sum()
        .rename(columns={"Load": "Total Load"})
    )
    role_month_df["Month Label"] = role_month_df["Month"].dt.strftime("%b-%Y")

    # Metrics for sorting/filtering roles.
    metrics_df = (
        role_month_df.groupby("Role", as_index=False)
        .agg(
            Max_Load=("Total Load", "max"),
            Avg_Load=("Total Load", "mean"),
            Active_Months=("Month", "count"),
        )
    )

    return detail_df, role_month_df, metrics_df


def add_overload_columns(
    detail_df: pd.DataFrame,
    role_month_df: pd.DataFrame,
    threshold: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    role_month_df = role_month_df.copy()
    role_month_df["Overload"] = role_month_df["Total Load"] > threshold
    role_month_df["Excess"] = (role_month_df["Total Load"] - threshold).clip(lower=0)
    role_month_df["Status"] = role_month_df["Overload"].map(
        {True: f"OVER > {threshold:g}", False: "OK"}
    )

    detail_df = detail_df.merge(
        role_month_df[["Role", "Month", "Total Load", "Overload", "Status"]],
        on=["Role", "Month"],
        how="left",
    )
    return detail_df, role_month_df


def build_timeline_segments(detail_df: pd.DataFrame) -> pd.DataFrame:
    """Convert monthly loading rows into continuous Gantt bars."""
    if detail_df.empty:
        return pd.DataFrame()

    segments = []
    group_cols = [
        "Project",
        "Activity No",
        "Activity",
        "Requirement",
        "Role",
        "Source Row",
        "Load",
        "Status",
        "Overload",
    ]

    for keys, group in detail_df.sort_values("Month").groupby(group_cols, dropna=False):
        group = group.sort_values("Month")
        current_start = None
        previous_month = None
        months_in_segment = []

        for month in group["Month"]:
            month = pd.Timestamp(month)
            if current_start is None:
                current_start = month
                previous_month = month
                months_in_segment = [month]
                continue

            expected_next = previous_month + pd.DateOffset(months=1)
            if month == expected_next:
                months_in_segment.append(month)
                previous_month = month
            else:
                end = previous_month + pd.DateOffset(months=1)
                segments.append(_segment_row(keys, group_cols, current_start, end, months_in_segment))
                current_start = month
                previous_month = month
                months_in_segment = [month]

        if current_start is not None:
            end = previous_month + pd.DateOffset(months=1)
            segments.append(_segment_row(keys, group_cols, current_start, end, months_in_segment))

    seg_df = pd.DataFrame(segments)
    if seg_df.empty:
        return seg_df

    seg_df["Load Label"] = seg_df["Load"].map(lambda x: f"{x:g}")
    seg_df["Activity Label"] = (
        seg_df["Role"]
        + " | "
        + seg_df["Project"]
        + " "
        + seg_df["Activity No"].astype(str)
        + " - "
        + seg_df["Activity"]
    )
    seg_df["Month Range"] = (
        seg_df["Start"].dt.strftime("%b-%Y")
        + " to "
        + (seg_df["Finish"] - pd.DateOffset(days=1)).dt.strftime("%b-%Y")
    )

    return seg_df


def _segment_row(keys, group_cols, start, finish, months):
    row = dict(zip(group_cols, keys))
    row["Start"] = start
    row["Finish"] = finish
    row["Months"] = len(months)
    return row


def get_download_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


# =========================
# FILE UPLOAD
# =========================
uploaded_file = st.file_uploader(
    "Upload Resource Planning workbook",
    type=["xlsx", "xlsm"],
)

if uploaded_file is None:
    st.info("Upload your Excel workbook to start.")
    st.stop()

try:
    detail_raw, role_month_raw, metrics_raw = parse_tc_activities(uploaded_file.getvalue())
except Exception as exc:
    st.error(f"Could not parse the workbook: {exc}")
    st.stop()


# =========================
# SIDEBAR FILTERS
# =========================
st.sidebar.header("Filters")

threshold = st.sidebar.number_input(
    "Overload threshold",
    min_value=0.1,
    value=1.0,
    step=0.1,
    help="A role is flagged as overloaded when total monthly load is above this value.",
)

projects = sorted(detail_raw["Project"].dropna().unique().tolist())
selected_projects = st.sidebar.multiselect(
    "Project section",
    projects,
    default=projects,
)

include_summary_rows = st.sidebar.checkbox(
    "Include summary/category rows",
    value=False,
    help="Keeps rows such as Total, ATC, ATS, SIG, Train, CSF, etc. Usually leave this off when checking individual person/role overload.",
)

# Apply project/summary filters first, then recalculate role-month totals.
# This makes the overload check respect the project sections selected in the sidebar.
base_detail = detail_raw[detail_raw["Project"].isin(selected_projects)].copy()

if not include_summary_rows:
    base_detail = base_detail[~base_detail["Role"].str.upper().isin(SUMMARY_ROLES)]

if base_detail.empty:
    st.warning("No manpower loading values found for the selected project filter.")
    st.stop()

base_role_month = (
    base_detail.groupby(["Role", "Month"], as_index=False)["Load"]
    .sum()
    .rename(columns={"Load": "Total Load"})
)
base_role_month["Month Label"] = base_role_month["Month"].dt.strftime("%b-%Y")

base_detail, base_role_month = add_overload_columns(
    base_detail,
    base_role_month,
    threshold,
)

filtered_detail = base_detail.copy()
filtered_role_month = base_role_month.copy()

role_metrics = (
    filtered_role_month.groupby("Role", as_index=False)
    .agg(
        Max_Load=("Total Load", "max"),
        Overloaded_Months=("Overload", "sum"),
        Active_Months=("Month", "count"),
    )
)

sort_option = st.sidebar.selectbox(
    "Sort role list by",
    ["Alphabetical", "Highest max load", "Most overloaded months"],
)

if sort_option == "Highest max load":
    role_options = role_metrics.sort_values(
        ["Max_Load", "Role"], ascending=[False, True]
    )["Role"].tolist()
elif sort_option == "Most overloaded months":
    role_options = role_metrics.sort_values(
        ["Overloaded_Months", "Max_Load", "Role"], ascending=[False, False, True]
    )["Role"].tolist()
else:
    role_options = sorted(role_metrics["Role"].tolist())

show_only_overloaded_roles = st.sidebar.checkbox(
    "Show only roles with overload",
    value=False,
)

if show_only_overloaded_roles:
    overloaded_roles = set(role_metrics.loc[role_metrics["Overloaded_Months"] > 0, "Role"])
    role_options = [role for role in role_options if role in overloaded_roles]

# Nice default: show overloaded roles first, capped so the chart remains readable.
default_roles = role_options[:8]
selected_roles = st.sidebar.multiselect(
    "Role / person",
    role_options,
    default=default_roles,
)

if not selected_roles:
    st.warning("Select at least one role/person from the sidebar.")
    st.stop()

filtered_detail = filtered_detail[filtered_detail["Role"].isin(selected_roles)].copy()
filtered_role_month = filtered_role_month[
    filtered_role_month["Role"].isin(selected_roles)
].copy()

min_month = filtered_detail["Month"].min().to_pydatetime()
max_month = filtered_detail["Month"].max().to_pydatetime()
selected_date_range = st.sidebar.slider(
    "Timeline range",
    min_value=min_month,
    max_value=max_month,
    value=(min_month, max_month),
    format="MMM YYYY",
)

start_filter = pd.Timestamp(selected_date_range[0])
end_filter = pd.Timestamp(selected_date_range[1])

filtered_detail = filtered_detail[
    (filtered_detail["Month"] >= start_filter) & (filtered_detail["Month"] <= end_filter)
].copy()
filtered_role_month = filtered_role_month[
    (filtered_role_month["Month"] >= start_filter)
    & (filtered_role_month["Month"] <= end_filter)
].copy()

annotate_bars = st.sidebar.checkbox("Annotate Gantt bars with load", value=True)


# =========================
# SUMMARY CARDS
# =========================
overload_rows = filtered_role_month[filtered_role_month["Overload"]].copy()
max_load = filtered_role_month["Total Load"].max() if not filtered_role_month.empty else 0
max_load_row = (
    filtered_role_month.loc[filtered_role_month["Total Load"].idxmax()]
    if not filtered_role_month.empty
    else None
)

col1, col2, col3, col4 = st.columns(4)
col1.metric("Selected roles", len(selected_roles))
col2.metric("Overloaded role-months", int(len(overload_rows)))
col3.metric("Highest monthly load", f"{max_load:.2f}")
if max_load_row is not None:
    col4.metric("Worst month", f"{max_load_row['Role']} / {max_load_row['Month Label']}")
else:
    col4.metric("Worst month", "-")

if len(overload_rows) > 0:
    st.error(
        f"⚠️ Overload detected: {len(overload_rows)} role-month(s) exceed {threshold:g}."
    )
else:
    st.success(f"No selected role/month exceeds {threshold:g}.")


# =========================
# TABS
# =========================
tab1, tab2, tab3, tab4 = st.tabs(
    ["Overload Heatmap", "Gantt Timeline", "Overload Details", "Parsed Data"]
)


# =========================
# TAB 1: HEATMAP
# =========================
with tab1:
    st.subheader("Monthly load by role")

    if filtered_role_month.empty:
        st.info("No monthly loading values found for the selected filters.")
    else:
        heatmap_df = filtered_role_month.pivot_table(
            index="Role",
            columns="Month",
            values="Total Load",
            aggfunc="sum",
            fill_value=0,
        )

        # Keep role order same as sidebar sort / selection.
        heatmap_df = heatmap_df.reindex([r for r in selected_roles if r in heatmap_df.index])
        heatmap_df.columns = [col.strftime("%b-%Y") for col in heatmap_df.columns]

        fig_heat = px.imshow(
            heatmap_df,
            aspect="auto",
            text_auto=".2g",
            color_continuous_scale="RdYlGn_r",
            labels=dict(x="Month", y="Role", color="Total Load"),
            title=f"Role load heatmap — red means closer to / above overload threshold {threshold:g}",
        )
        fig_heat.update_layout(height=max(450, 38 * len(heatmap_df)))
        fig_heat.update_xaxes(side="top", tickangle=45)
        st.plotly_chart(fig_heat, use_container_width=True)


# =========================
# TAB 2: GANTT
# =========================
with tab2:
    st.subheader("Activity timeline by selected role")

    segment_df = build_timeline_segments(filtered_detail)

    if segment_df.empty:
        st.info("No Gantt bars found for the selected filters.")
    else:
        gantt_sort = st.radio(
            "Gantt row order",
            ["By role then start date", "By project/activity", "By start date"],
            horizontal=True,
        )

        if gantt_sort == "By project/activity":
            segment_df = segment_df.sort_values(
                ["Project", "Activity No", "Role", "Start", "Load"]
            )
        elif gantt_sort == "By start date":
            segment_df = segment_df.sort_values(["Start", "Role", "Project", "Activity No"])
        else:
            role_order_map = {role: i for i, role in enumerate(selected_roles)}
            segment_df["Role Order"] = segment_df["Role"].map(role_order_map)
            segment_df = segment_df.sort_values(
                ["Role Order", "Start", "Project", "Activity No", "Activity"]
            )

        text_col = "Load Label" if annotate_bars else None

        fig_gantt = px.timeline(
            segment_df,
            x_start="Start",
            x_end="Finish",
            y="Activity Label",
            color="Status",
            text=text_col,
            color_discrete_map={
                "OK": "#2E7D32",
                f"OVER > {threshold:g}": "#C62828",
            },
            hover_data={
                "Role": True,
                "Project": True,
                "Activity No": True,
                "Activity": True,
                "Requirement": True,
                "Load": ":.2f",
                "Status": True,
                "Month Range": True,
                "Activity Label": False,
                "Start": False,
                "Finish": False,
            },
            title="Gantt bars are split when the load or overload status changes",
        )
        fig_gantt.update_yaxes(autorange="reversed")
        fig_gantt.update_xaxes(dtick="M1", tickformat="%b\n%Y", side="top")
        fig_gantt.update_layout(
            height=max(500, 28 * len(segment_df["Activity Label"].unique())),
            legend_title_text="Status",
        )
        st.plotly_chart(fig_gantt, use_container_width=True)


# =========================
# TAB 3: OVERLOAD DETAILS
# =========================
with tab3:
    st.subheader("Exact months where selected roles exceed threshold")

    if overload_rows.empty:
        st.success("No overload rows for the selected filters.")
    else:
        contributors = (
            filtered_detail[filtered_detail["Overload"]]
            .groupby(["Role", "Month", "Month Label"], as_index=False)
            .agg(
                Total_Load=("Total Load", "max"),
                Activities=(
                    "Activity",
                    lambda x: " | ".join(sorted(set(str(v) for v in x if str(v) != "nan"))),
                ),
                Projects=("Project", lambda x: ", ".join(sorted(set(x)))),
            )
        )
        contributors["Excess"] = contributors["Total_Load"] - threshold
        contributors = contributors.sort_values(
            ["Total_Load", "Role", "Month"], ascending=[False, True, True]
        )

        display_cols = [
            "Role",
            "Month Label",
            "Total_Load",
            "Excess",
            "Projects",
            "Activities",
        ]
        st.dataframe(
            contributors[display_cols],
            use_container_width=True,
            hide_index=True,
        )
        st.download_button(
            "Download overload details CSV",
            data=get_download_csv(contributors[display_cols]),
            file_name="overload_details.csv",
            mime="text/csv",
        )


# =========================
# TAB 4: PARSED DATA
# =========================
with tab4:
    st.subheader("Parsed monthly records")
    st.write(
        "This is the cleaned long-format data extracted from `T&C Activities`. "
        "Each row is one role/activity/month loading value."
    )

    display_detail = filtered_detail.sort_values(
        ["Role", "Month", "Project", "Activity No", "Activity"]
    )[
        [
            "Project",
            "Activity No",
            "Activity",
            "Requirement",
            "Role",
            "Month Label",
            "Load",
            "Total Load",
            "Status",
            "Source Row",
        ]
    ]

    st.dataframe(display_detail, use_container_width=True, hide_index=True)
    st.download_button(
        "Download parsed data CSV",
        data=get_download_csv(display_detail),
        file_name="tc_activities_parsed_data.csv",
        mime="text/csv",
    )

